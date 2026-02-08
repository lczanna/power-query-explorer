"""
Generate test Excel (.xlsx) files with real, Excel-visible Power Query metadata.

Why this script uses a template:
- A workbook that shows queries in Excel requires more than just DataMashup.
- It also needs related connection/query table/customXml parts.
- The template file contains that valid structure, and this script only swaps
  Formulas/Section1.m inside DataMashup.

Files are saved to data/test-files/.
"""

import base64
import io
import os
import shutil
import struct
import zipfile
from xml.etree import ElementTree as ET

import openpyxl

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(ROOT_DIR, "data", "test-files")
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates", "query-template.xlsx")

DATAMASHUP_XML_PATH = "customXml/item1.xml"
DATAMASHUP_NS = "http://schemas.microsoft.com/DataMashup"


def _decode_xml(raw: bytes) -> str:
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return raw.decode("utf-16")
    return raw.decode("utf-8")


def _read_datamashup_blob(item1_xml: bytes):
    xml_text = _decode_xml(item1_xml)
    root = ET.fromstring(xml_text)
    blob = base64.b64decode((root.text or "").strip())

    if len(blob) < 8:
        raise ValueError("DataMashup blob is too short")

    version, pkg_len = struct.unpack("<II", blob[:8])
    if len(blob) < 8 + pkg_len:
        raise ValueError("DataMashup package length is invalid")

    package = blob[8 : 8 + pkg_len]
    trailer = blob[8 + pkg_len :]
    return root, version, package, trailer


def _replace_section1_m(package_bytes: bytes, m_code: str) -> bytes:
    in_buf = io.BytesIO(package_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        names = zin.namelist()
        parts = {name: zin.read(name) for name in names}

    parts["Formulas/Section1.m"] = m_code.encode("utf-8")

    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in names:
            if name in parts and name not in written:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)

    return out_buf.getvalue()


def _build_datamashup_xml(root: ET.Element, version: int, package: bytes, trailer: bytes) -> bytes:
    ET.register_namespace("", DATAMASHUP_NS)
    blob = struct.pack("<I", version) + struct.pack("<I", len(package)) + package + trailer
    root.text = base64.b64encode(blob).decode("ascii")
    return ET.tostring(root, encoding="utf-16", xml_declaration=True)


def create_query_workbook(path: str, m_code: str) -> None:
    """Create a query-enabled workbook from template and replace Section1.m."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Missing template workbook: {TEMPLATE_PATH}")

    shutil.copy2(TEMPLATE_PATH, path)

    with zipfile.ZipFile(path, "r") as zin:
        order = [name for name in zin.namelist() if not name.endswith("/")]
        parts = {name: zin.read(name) for name in order}

    root, version, package, trailer = _read_datamashup_blob(parts[DATAMASHUP_XML_PATH])
    package = _replace_section1_m(package, m_code)
    parts[DATAMASHUP_XML_PATH] = _build_datamashup_xml(root, version, package, trailer)

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in order:
            if name in parts and name not in written:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)

    with open(path, "wb") as f:
        f.write(out_buf.getvalue())


# ============================================================
# M Code Definitions
# ============================================================


def m_simple_query() -> str:
    return (
        "section Section1;\n"
        "shared #\"Query1\" = let\n"
        '    Source = #table({"OrderID", "CustomerID", "Amount", "OrderDate", "Category"}, {{1001, 1, 240.5, #date(2024, 1, 4), "Hardware"}, {1002, 2, 180.0, #date(2024, 1, 5), "Services"}, {1003, 1, 95.0, #date(2024, 1, 10), "Returns"}, {1004, 3, 420.9, #date(2024, 1, 12), "Hardware"}, {1005, 4, 75.2, #date(2024, 1, 13), "Returns"}, {1006, 2, 330.0, #date(2024, 2, 1), "Services"}, {1007, 5, 510.4, #date(2024, 2, 6), "Hardware"}, {1008, 4, 125.5, #date(2024, 2, 10), "Accessories"}, {1009, 3, 290.0, #date(2024, 2, 12), "Hardware"}, {1010, 6, 60.0, #date(2024, 2, 14), "Returns"}, {1011, 5, 700.3, #date(2024, 2, 21), "Hardware"}, {1012, 7, 155.8, #date(2024, 2, 25), "Services"}})\n'
        "in\n"
        "    Source;\n"
        "shared SalesData = let\n"
        "    Source = Query1,\n"
        '    ChangedType = Table.TransformColumnTypes(Source, {{"Amount", type number}, {"OrderDate", type date}, {"Category", type text}}),\n'
        '    FilteredRows = Table.SelectRows(ChangedType, each [Amount] > 100 and [Category] <> "Returns"),\n'
        '    SortedRows = Table.Sort(FilteredRows, {{"OrderDate", Order.Ascending}})\n'
        "in\n"
        "    SortedRows;\n"
        "shared EdgeLookup = let\n"
        '    Source = Excel.Workbook(File.Contents("edge_cases.xlsx"), null, true)\n'
        "in\n"
        "    Source;\n"
    )


def m_multi_query() -> str:
    return (
        "section Section1;\n"
        "shared #\"Query1\" = let\n"
        '    Source = Excel.Workbook(File.Contents("simple_query.xlsx"), null, true)\n'
        "in\n"
        "    Source;\n"
        "shared RawOrders = let\n"
        '    Source = #table({"OrderID", "CustomerID", "Amount", "Priority"}, {{2001, 1, 240.5, "High"}, {2002, 2, 180.0, "Low"}, {2003, 1, 95.0, "Low"}, {2004, 3, 420.9, "High"}, {2005, 4, 75.2, "Low"}, {2006, 2, 330.0, "High"}, {2007, 5, 510.4, "High"}, {2008, 4, 125.5, "Low"}, {2009, 3, 290.0, "Medium"}, {2010, 6, 60.0, "Low"}, {2011, 5, 700.3, "High"}, {2012, 7, 155.8, "Medium"}}),\n'
        '    ChangedType = Table.TransformColumnTypes(Source, {{"OrderID", Int64.Type}, {"CustomerID", Int64.Type}, {"Amount", type number}, {"Priority", type text}})\n'
        "in\n"
        "    ChangedType;\n"
        "shared Customers = let\n"
        '    Source = #table({"CustomerID", "Name", "Region", "Segment"}, {{1, "Acme Corp", "West", "Enterprise"}, {2, "Bluebird LLC", "East", "SMB"}, {3, "Northwind", "North", "Enterprise"}, {4, "Fabrikam", "South", "Mid-Market"}, {5, "Contoso", "West", "Enterprise"}, {6, "Adventure Works", "East", "SMB"}, {7, "Tailspin Toys", "North", "SMB"}}),\n'
        '    ChangedType = Table.TransformColumnTypes(Source, {{"CustomerID", Int64.Type}, {"Name", type text}, {"Region", type text}, {"Segment", type text}})\n'
        "in\n"
        "    ChangedType;\n"
        "shared OrdersWithCustomers = let\n"
        "    Source = RawOrders,\n"
        '    Merged = Table.NestedJoin(Source, {"CustomerID"}, Customers, {"CustomerID"}, "CustomerData", JoinKind.LeftOuter),\n'
        '    Expanded = Table.ExpandTableColumn(Merged, "CustomerData", {"Name", "Region", "Segment"})\n'
        "in\n"
        "    Expanded;\n"
        "shared SalesSummary = let\n"
        "    Source = OrdersWithCustomers,\n"
        '    Grouped = Table.Group(Source, {"Region"}, {{"TotalSales", each List.Sum([Amount]), type number}, {"OrderCount", each Table.RowCount(_), Int64.Type}, {"AvgOrder", each Number.Round(List.Average([Amount]), 2), type number}}),\n'
        '    Sorted = Table.Sort(Grouped, {{"TotalSales", Order.Descending}})\n'
        "in\n"
        "    Sorted;\n"
    )


def m_complex_code() -> str:
    return (
        "section Section1;\n"
        "shared #\"Query1\" = let\n"
        '    Source = #table({"SalesOrderID", "CustomerID", "TerritoryID", "Status", "TotalDue", "OrderDate"}, {{5001, 1, 1, "Open", 125.3, #date(2023, 1, 7)}, {5002, 2, 4, "Open", 840.0, #date(2023, 1, 9)}, {5003, 3, 8, "Cancelled", 400.5, #date(2023, 1, 12)}, {5004, 2, 11, "Open", 99.9, #date(2023, 2, 2)}, {5005, 4, 6, "Closed", 1520.2, #date(2023, 2, 9)}, {5006, 5, 2, null, 730.0, #date(2023, 2, 11)}, {5007, 6, 9, "Open", 0.0, #date(2023, 3, 1)}, {5008, 1, 3, "Closed", 275.8, #date(2023, 3, 5)}, {5009, 7, 7, "Open", 319.4, #date(2023, 3, 9)}, {5010, 8, 10, "Open", 450.2, #date(2023, 3, 11)}}),\n'
        '    ChangedType = Table.TransformColumnTypes(Source, {{"SalesOrderID", Int64.Type}, {"CustomerID", Int64.Type}, {"TerritoryID", Int64.Type}, {"Status", type text}, {"TotalDue", type number}, {"OrderDate", type date}})\n'
        "in\n"
        "    ChangedType;\n"
        "shared ExternalBaseline = let\n"
        '    Source = Excel.Workbook(File.Contents("multi_query.xlsx"), null, true)\n'
        "in\n"
        "    Source;\n"
        "shared TransformPipeline = let\n"
        "    // Load data from in-workbook query\n"
        "    Source = Query1,\n"
        "\n"
        "    /* Multi-line comment:\n"
        "       This step filters for active orders\n"
        "       and handles null values */\n"
        "    FilteredRows = Table.SelectRows(Source, each\n"
        "        [Status] <> null\n"
        '        and [Status] <> "Cancelled"\n'
        "        and [TotalDue] > 0\n"
        "    ),\n"
        "\n"
        "    // Nested let expression for custom transformation\n"
        "    AddedColumns = let\n"
        "        withTax = Table.AddColumn(FilteredRows, \"TaxRate\", each\n"
        "            if [TerritoryID] >= 1 and [TerritoryID] <= 5 then 0.08\n"
        "            else if [TerritoryID] >= 6 and [TerritoryID] <= 10 then 0.065\n"
        "            else 0.05\n"
        "        ),\n"
        '        withTotal = Table.AddColumn(withTax, "GrossTotal", each [TotalDue] * (1 + [TaxRate]), type number)\n'
        "    in\n"
        "        withTotal,\n"
        "\n"
        "    // Handle special characters in column renaming\n"
        '    RenamedColumns = Table.RenameColumns(AddedColumns, {{"SalesOrderID", "Order #"}, {"CustomerID", "Cust_ID"}}),\n'
        "\n"
        "    // Error handling with try/otherwise\n"
        "    SafeConversion = Table.TransformColumns(RenamedColumns, {\n"
        '        {"GrossTotal", each try Number.Round(_, 2) otherwise 0, type number}\n'
        "    }),\n"
        "\n"
        "    // Dynamic parameter usage\n"
        '    ParamFiltered = if #"Start Date" <> null\n'
        '        then Table.SelectRows(SafeConversion, each [OrderDate] >= #"Start Date")\n'
        "        else SafeConversion,\n"
        "\n"
        "    Result = Table.Buffer(ParamFiltered)\n"
        "in\n"
        "    Result;\n"
        'shared #"Start Date" = let\n'
        '    Source = #date(2023, 1, 1)\n'
        "in\n"
        "    Source;\n"
    )


def m_stress_test() -> str:
    queries = [
        (
            "shared #\"Query1\" = let\n"
            '    Source = Excel.Workbook(File.Contents("complex_code.xlsx"), null, true)\n'
            "in\n"
            "    Source"
        )
    ]

    categories = ["Hardware", "Software", "Services", "Support"]
    for i in range(1, 6):
        rows = []
        for row_id in range(1, 61):
            cat = categories[(row_id + i) % len(categories)]
            rows.append(f"{{{row_id}, {row_id * (i + 4) + (i * 3)}, \"{cat}\"}}")
        table_literal = (
            '#table({"ID", "Value", "Category"}, {' + ", ".join(rows) + "})"
        )
        queries.append(
            f"shared DataSource{i} = let\n"
            f"    Source = {table_literal},\n"
            f'    ChangedType = Table.TransformColumnTypes(Source, {{ {{"ID", Int64.Type}}, {{"Value", type number}}, {{"Category", type text}} }})\n'
            f"in\n"
            f"    ChangedType"
        )

    for i in range(1, 6):
        queries.append(
            f"shared Transform{i} = let\n"
            f"    Source = DataSource{i},\n"
            f"    Filtered = Table.SelectRows(Source, each [Value] > {i * 10}),\n"
            f'    Added = Table.AddColumn(Filtered, "Computed", each [Value] * {i})\n'
            f"in\n"
            f"    Added"
        )

    for i in range(1, 6):
        j = (i % 5) + 1
        queries.append(
            f"shared Merge{i} = let\n"
            f"    Left = Transform{i},\n"
            f"    Right = Transform{j},\n"
            f'    Merged = Table.NestedJoin(Left, {{"ID"}}, Right, {{"ID"}}, "Joined", JoinKind.Inner),\n'
            f'    Expanded = Table.ExpandTableColumn(Merged, "Joined", {{"Value", "Computed"}})\n'
            f"in\n"
            f"    Expanded"
        )

    for i in range(1, 6):
        queries.append(
            f"shared Aggregate{i} = let\n"
            f"    Source = Merge{i},\n"
            f'    Grouped = Table.Group(Source, {{"Category"}}, {{ {{"Total", each List.Sum([Value]), type number}}, {{"Count", each Table.RowCount(_), Int64.Type}} }}),\n'
            f'    Sorted = Table.Sort(Grouped, {{ {{"Total", Order.Descending}} }})\n'
            f"in\n"
            f"    Sorted"
        )

    for i in range(1, 6):
        queries.append(
            f"shared Output{i} = let\n"
            f"    Source = Aggregate{i},\n"
            f"    TopN = Table.FirstN(Source, 10),\n"
            f'    Renamed = Table.RenameColumns(TopN, {{ {{"Total", "Grand Total {i}"}}, {{"Count", "Record Count {i}"}} }})\n'
            f"in\n"
            f"    Renamed"
        )

    return "section Section1;\n" + ";\n".join(queries) + ";"


def m_edge_cases() -> str:
    long_steps = "".join(
        [
            f'    Step{i} = Table.AddColumn({"Source" if i == 1 else f"Step{i-1}"}, "Col{i}", each [Units] * {i}),\\n'
            for i in range(1, 21)
        ]
    )

    return (
        "section Section1;\n"
        "shared #\"Query1\" = let\n"
        '    Source = Excel.Workbook(File.Contents("simple_query.xlsx"), null, true)\n'
        "in\n"
        "    Source;\n"
        'shared #"Revenue Report" = let\n'
        '    Source = #table({"Product", "Revenue", "Quarter"}, {{"Widget A", 15000, "Q1"}, {"Widget B", 23100, "Q1"}, {"Widget C", 19850, "Q2"}, {"Widget A", 17600, "Q2"}, {"Widget D", 26400, "Q3"}, {"Widget B", 24550, "Q3"}})\n'
        "in\n"
        "    Source;\n"
        "shared raw_data_import = let\n"
        '    Source = #table({"Year", "Month", "Units", "Region"}, {{2024, 1, 18, "West"}, {2024, 2, 24, "East"}, {2024, 3, 31, "North"}, {2025, 1, 22, "West"}, {2025, 2, 29, "South"}, {2025, 3, 35, "East"}}),\n'
        '    Headers = Table.TransformColumnTypes(Source, {{"Year", Int64.Type}, {"Month", Int64.Type}, {"Units", Int64.Type}, {"Region", type text}})\n'
        "in\n"
        "    Headers;\n"
        'shared #"Year-to-Date (YTD) #1" = let\n'
        "    Source = raw_data_import,\n"
        "    Filtered = Table.SelectRows(Source, each [Year] = Date.Year(DateTime.LocalNow()))\n"
        "in\n"
        "    Filtered;\n"
        "shared EmptyResult = let\n"
        "    Source = null\n"
        "in\n"
        "    Source;\n"
        "shared LongProcessing = let\n"
        "    Source = raw_data_import,\n"
        + long_steps
        + "    Final = Table.Buffer(Step20)\n"
        "in\n"
        "    Final;\n"
        'shared #"Donn\u00e9es_brutes" = let\n'
        '    Source = #table({"Cle", "Montant", "Commentaire"}, {{"A1", 10.5, "Ligne initiale"}, {"B2", 22.0, "Donn\u00e9es test"}, {"C3", 35.75, "Valeur moyenne"}, {"D4", 41.2, "Fin de s\u00e9rie"}})\n'
        "in\n"
        "    Source;\n"
    )


def create_pbix_file(path: str, m_code: str) -> None:
    """Create a minimal .pbix/.pbit file with a DataMashup containing the given M code.

    A .pbix file is a ZIP (OPC) archive. The Power Query M code lives inside a
    DataMashup binary (MS-QDEFF format) at the root of the archive. We build
    a minimal archive with just enough structure for the explorer to parse.
    """
    # Build inner OPC ZIP (the Package inside DataMashup)
    inner_buf = io.BytesIO()
    with zipfile.ZipFile(inner_buf, "w", zipfile.ZIP_DEFLATED) as zin:
        zin.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="utf-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Default Extension="m" ContentType="application/x-ms-m"/>'
            "</Types>",
        )
        zin.writestr(
            "Config/Package.xml",
            '<?xml version="1.0" encoding="utf-8"?>'
            '<Package xmlns="http://schemas.microsoft.com/DataMashup">'
            "<Version>2.72.0</Version>"
            "<MinVersion>2.21.0</MinVersion>"
            "<Culture>en-US</Culture>"
            "</Package>",
        )
        zin.writestr("Formulas/Section1.m", m_code)
    inner_bytes = inner_buf.getvalue()

    # Build MS-QDEFF binary stream: version(4) + pkgLen(4) + ZIP(pkgLen)
    version = 0
    datamashup_blob = struct.pack("<II", version, len(inner_bytes)) + inner_bytes

    # Build the outer .pbix ZIP archive
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        zout.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="utf-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="json" ContentType="application/json"/>'
            '<Override PartName="/DataMashup" ContentType="application/vnd.ms-DataMashup"/>'
            "</Types>",
        )
        zout.writestr("DataMashup", datamashup_blob)
        zout.writestr("Version", "1.0")
        zout.writestr("DataModelSchema", "{}")
        zout.writestr(
            "Report/Layout",
            '{"id":0,"reportId":"00000000-0000-0000-0000-000000000000","pods":[]}',
        )
        zout.writestr("Settings", '{"version":"1.0"}')
        zout.writestr("Metadata", '{"version":"1.0","type":"report"}')

    with open(path, "wb") as f:
        f.write(out_buf.getvalue())


# ============================================================
# File Creators
# ============================================================


def create_simple_query():
    path = os.path.join(OUTPUT_DIR, "simple_query.xlsx")
    create_query_workbook(path, m_simple_query())
    print(f"  Created: {path}")


def create_multi_query():
    path = os.path.join(OUTPUT_DIR, "multi_query.xlsx")
    create_query_workbook(path, m_multi_query())
    print(f"  Created: {path}")


def create_complex_code():
    path = os.path.join(OUTPUT_DIR, "complex_code.xlsx")
    create_query_workbook(path, m_complex_code())
    print(f"  Created: {path}")


def create_stress_test():
    path = os.path.join(OUTPUT_DIR, "stress_test.xlsx")
    create_query_workbook(path, m_stress_test())
    print(f"  Created: {path}")


def create_no_queries():
    path = os.path.join(OUTPUT_DIR, "no_queries.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    for i in range(2, 12):
        ws[f"A{i}"] = f"Item {i - 1}"
        ws[f"B{i}"] = i * 10
    wb.save(path)
    print(f"  Created: {path}")


def create_edge_cases():
    path = os.path.join(OUTPUT_DIR, "edge_cases.xlsx")
    create_query_workbook(path, m_edge_cases())
    print(f"  Created: {path}")


def create_simple_pbix():
    path = os.path.join(OUTPUT_DIR, "simple_query.pbix")
    create_pbix_file(path, m_simple_query())
    print(f"  Created: {path}")


def create_multi_pbix():
    path = os.path.join(OUTPUT_DIR, "multi_query.pbix")
    create_pbix_file(path, m_multi_query())
    print(f"  Created: {path}")


def create_simple_pbit():
    path = os.path.join(OUTPUT_DIR, "simple_query.pbit")
    create_pbix_file(path, m_simple_query())
    print(f"  Created: {path}")


# ============================================================
# Main
# ============================================================


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Creating test files...")
    create_simple_query()
    create_multi_query()
    create_complex_code()
    create_stress_test()
    create_no_queries()
    create_edge_cases()
    create_simple_pbix()
    create_multi_pbix()
    create_simple_pbit()
    print(f"\nAll test files created in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
